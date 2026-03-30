import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from constants import (
    CII_CODES, NON_CII_CODES, ALL_CODES,
    IMMO_CODES, CHARGES_CODES, ALL_COMPTABLE_CODES, COMPTABLE_DETAILS,
    PROJECT_CODES, MONTHS,
    CII_DETAILS, NON_CII_DETAILS, PROJECT_DETAILS,
    DISPLAY_COLUMNS, COMPTABLE_DISPLAY,
)

SOURCE_FILE = "gitlab_classified.xlsx"
OUTPUT_FILE = "tableaux_CII_mensuels_realistes.xlsx"


def load_events() -> pd.DataFrame:
    events = pd.read_excel(SOURCE_FILE, sheet_name="events")
    events["event_date"] = pd.to_datetime(events["event_date"], errors="coerce")
    events["date"] = events["event_date"].dt.date

    # Validation des codes saisis manuellement
    if "cii_code" in events.columns:
        invalid_cii = events["cii_code"].dropna()
        invalid_cii = invalid_cii[~invalid_cii.isin(ALL_CODES)]
        if not invalid_cii.empty:
            print(f"⚠️  {len(invalid_cii)} événements avec cii_code invalide : {invalid_cii.unique().tolist()}")

    if "comptable" in events.columns:
        invalid_compta = events["comptable"].dropna()
        invalid_compta = invalid_compta[~invalid_compta.isin(ALL_COMPTABLE_CODES)]
        if not invalid_compta.empty:
            print(f"⚠️  {len(invalid_compta)} événements avec comptable invalide : {invalid_compta.unique().tolist()}")

    if "project" in events.columns:
        invalid_proj = events["project"].dropna()
        invalid_proj = invalid_proj[~invalid_proj.isin(PROJECT_CODES)]
        if not invalid_proj.empty:
            print(f"⚠️  {len(invalid_proj)} événements avec project invalide : {invalid_proj.unique().tolist()}")

    return events


def build_daily_allocations(events: pd.DataFrame) -> pd.DataFrame:
    rows = []
    code_defaults = {c: 0.0 for c in ALL_CODES}
    comptable_defaults = {c: 0.0 for c in ALL_COMPTABLE_CODES}
    proj_defaults = {p: 0.0 for p in PROJECT_CODES}

    for (person_key, display_name, date_value, month_name), g in events.groupby(
        ["person_key", "display_name", "date", "month_name"], sort=True
    ):
        day_value = min(1.0, g["weight"].sum())
        if day_value <= 0:
            continue

        total_weight = g["weight"].sum()
        code_amounts = dict(code_defaults)
        comptable_amounts = dict(comptable_defaults)
        proj_amounts = dict(proj_defaults)

        for _, r in g.iterrows():
            allocated = day_value * (r["weight"] / total_weight) if total_weight else 0.0
            if r["cii_code"] in code_amounts:
                code_amounts[r["cii_code"]] += allocated
            if r.get("comptable") in comptable_amounts:
                comptable_amounts[r["comptable"]] += allocated
            if r["project"] in proj_amounts:
                proj_amounts[r["project"]] += allocated

        row = {
            "person_key": person_key,
            "username": display_name,
            "date": pd.to_datetime(date_value),
            "month": month_name,
            "TOTAL": round(day_value, 6),
        }
        row.update(code_amounts)
        row["TOTAL CII"] = sum(code_amounts.get(c, 0) for c in CII_CODES)
        row["TOTAL HORS CII"] = sum(code_amounts.get(c, 0) for c in NON_CII_CODES)
        row.update(comptable_amounts)
        row["TOTAL IMMO"] = sum(comptable_amounts.get(c, 0) for c in IMMO_CODES)
        row["TOTAL CHARGES"] = sum(comptable_amounts.get(c, 0) for c in CHARGES_CODES)
        row.update(proj_amounts)
        rows.append(row)

    return pd.DataFrame(rows)


def build_month_sheet(daily: pd.DataFrame, month_name: str, users: list[str], columns: list[str]) -> pd.DataFrame:
    month_df = daily[daily["month"] == month_name].copy()
    grouped = month_df.groupby("username")[columns].sum()
    grouped = grouped.reindex(users, fill_value=0.0)
    grouped = grouped.round(1)
    grouped.index.name = "username"
    grouped.loc["TOTAL"] = grouped.sum(numeric_only=True)
    return grouped


def build_resources_sheet(daily: pd.DataFrame, users: list[str]) -> pd.DataFrame:
    reverse_month = {v: k for k, v in MONTHS}
    rows = []
    for user in users:
        g = daily[daily["username"] == user]
        if g.empty:
            continue
        months = sorted({reverse_month[m] for m in g["month"].dropna().tolist()})
        cii_total = g[CII_CODES].sum().sum()
        immo_total = g[IMMO_CODES].sum().sum()
        total = g["TOTAL"].sum()
        rows.append({
            "Nom": user,
            "Mois début": min(months),
            "Mois fin": max(months),
            "Mois actifs": len(months),
            "Jours CII": round(cii_total, 1),
            "Jours hors CII": round(total - cii_total, 1),
            "Jours immobilisables": round(immo_total, 1),
            "Jours charges": round(total - immo_total, 1),
            "Jours total": round(total, 1),
        })
    return pd.DataFrame(rows).sort_values("Nom")


def build_legend_sheet() -> pd.DataFrame:
    rows = []
    for c in CII_CODES:
        rows.append({"Type": "CII", "Code": c, "Description": CII_DETAILS[c]})
    for c in NON_CII_CODES:
        rows.append({"Type": "HORS CII", "Code": c, "Description": NON_CII_DETAILS[c]})
    for c in ALL_COMPTABLE_CODES:
        immo = "Immobilisable" if c in IMMO_CODES else "Charges"
        rows.append({"Type": f"COMPTABLE ({immo})", "Code": c, "Description": COMPTABLE_DETAILS[c]})
    for p in PROJECT_CODES:
        rows.append({"Type": "PROJET", "Code": p, "Description": PROJECT_DETAILS[p]})
    rows.append({
        "Type": "RÈGLE",
        "Code": "TOTAL",
        "Description": "Somme réelle des jours journaliers calculés, avec plafond de 1,0 jour par personne et par date.",
    })
    return pd.DataFrame(rows)


def apply_annual_formulas(path: str, sheet_name: str, users: list[str], columns: list[str]) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    month_names = [m for _, m in MONTHS]

    for r, user in enumerate(users + ["TOTAL"], start=2):
        ws.cell(row=r, column=1, value=user)
        for c in range(2, len(columns) + 2):
            col = get_column_letter(c)
            # Référencer les feuilles mensuelles correspondantes (CII ou IMMO)
            prefix = "CII " if sheet_name == "Annuel CII" else "IMMO " if sheet_name == "Annuel IMMO" else ""
            refs = [f"'{prefix}{m}'!{col}{r}" for m in month_names]
            ws.cell(row=r, column=c, value="=" + "+".join(refs))

    wb.save(path)


def main():
    events = load_events()
    daily = build_daily_allocations(events)
    users = sorted(daily["username"].dropna().unique().tolist())

    legend_df = build_legend_sheet()
    resources_df = build_resources_sheet(daily, users)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Annuels
        pd.DataFrame(index=users + ["TOTAL"], columns=DISPLAY_COLUMNS).to_excel(
            writer, sheet_name="Annuel CII")
        pd.DataFrame(index=users + ["TOTAL"], columns=COMPTABLE_DISPLAY).to_excel(
            writer, sheet_name="Annuel IMMO")

        legend_df.to_excel(writer, sheet_name="Légende", index=False)
        resources_df.to_excel(writer, sheet_name="Ressources", index=False)
        daily.to_excel(writer, sheet_name="Détail journalier", index=False)

        # Feuilles mensuelles CII
        for _, month_name in MONTHS:
            build_month_sheet(daily, month_name, users, DISPLAY_COLUMNS).to_excel(
                writer, sheet_name=f"CII {month_name}")

        # Feuilles mensuelles IMMO
        for _, month_name in MONTHS:
            build_month_sheet(daily, month_name, users, COMPTABLE_DISPLAY).to_excel(
                writer, sheet_name=f"IMMO {month_name}")

    apply_annual_formulas(OUTPUT_FILE, "Annuel CII", users, DISPLAY_COLUMNS)
    apply_annual_formulas(OUTPUT_FILE, "Annuel IMMO", users, COMPTABLE_DISPLAY)
    print("✅ Fichier généré avec double ventilation CII + Immobilisation.")


if __name__ == "__main__":
    main()
